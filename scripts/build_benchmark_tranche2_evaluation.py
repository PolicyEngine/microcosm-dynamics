#!/usr/bin/env python3
"""Build the audited tranche-2 benchmark evaluation artifact.

The builder is deliberately independent of the standing registry.  It verifies
every audited capture and committed model artifact before parsing any of them,
then emits the 60 inventory actions used to evaluate the expanded registry.
The artifact is validation-only and never runs or estimates the model.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import statistics
from collections import Counter, defaultdict
from copy import deepcopy
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "runs/benchmark_tranche2_evaluation_v1.json"
VALIDATION_SOURCES = Path.home() / "PolicyEngine/psid-data/validation-sources"
VALIDATION_MANIFEST = VALIDATION_SOURCES / "manifest.jsonl"
VALIDATION_MANIFEST_SHA256 = (
    "72c180e8d162d9cc09017c355214ba0f9e1175b2d79f294ec2de96ee28cb2e1a"
)
VALIDATION_MANIFEST_SIZE_BYTES = 12_042

SSA_SNAPSHOT_DIR = ROOT / "data/external/snapshots/ssa_level_anchors_vintage1"
SSA_CAPTURE_MANIFEST = SSA_SNAPSHOT_DIR / "capture_manifest.txt"
SSA_CAPTURE_MANIFEST_SHA256 = (
    "569dbed5922c2192277eb671685a5859ba1440e30289e8c252e1762956c150ca"
)
SSA_CAPTURE_MANIFEST_SIZE_BYTES = 698
SSA_CAPTURE = SSA_SNAPSHOT_DIR / "supplement2025_4b.html"
SSA_CAPTURE_SHA256 = (
    "c228920ea9d53b1e323e5933b6d9f926e3c9b609d868b549fabc40118554b449"
)
SSA_CAPTURE_SIZE_BYTES = 488_165
SSA_EXTRACTION = (
    ROOT / "data/external/ssa_supplement2025_4b7_band_shares_v1.json"
)
SSA_EXTRACTION_SHA256 = (
    "c856acfc7d5fad8cb4143134f10fa48471255c87848194276ac8a9c46ec07bcb"
)
SSA_EXTRACTION_SIZE_BYTES = 14_840

CAPTURE_SPECS = {
    "cbo_55038_supplemental": {
        "filename": "cbo-att-55038-SupplementalData.xlsx",
        "sha256": (
            "9403fe44c44b360276d9ccf21a85ef5a55d4e689020ceab715a9e33e38d8429c"
        ),
        "size_bytes": 6_147_155,
    },
    "cbo_60392_data": {
        "filename": "cbo-60392-Data.xlsx",
        "sha256": (
            "1d66e714a8508f698e2528c7d295d76ab4a5659d91fc6cfc685230c536f30607"
        ),
        "size_bytes": 2_556_776,
    },
    "cbo_60392_long_term": {
        "filename": "cbo-60392-Long-Term-Projections.xlsx",
        "sha256": (
            "8945d7c5599e944e5786801daf3af9ebad318b24b871198708eeff9bb6c46f7b"
        ),
        "size_bytes": 85_665,
    },
    "cbo_60392_additional": {
        "filename": "cbo-60392-Additional-Info.xlsx",
        "sha256": (
            "9ff92c8e54e5b873f4b7743e695773876714a23c425c3ccb0c5dbb8d0c4dc739"
        ),
        "size_bytes": 67_661,
    },
    "mint_beneficiaries": {
        "filename": "ssa-mint-tables-beneficiaries.html",
        "sha256": (
            "a3f6da991356045f165b6517390ab8fbdd84e1de386e2d026ab5dc2ff9009ccf"
        ),
        "size_bytes": 1_652_291,
    },
    "mint_taxpayers": {
        "filename": "ssa-mint-tables-taxpayers.html",
        "sha256": (
            "0ee73c8c07639d69b29a69297028b6db12799115579c7ae3b663a8856031152e"
        ),
        "size_bytes": 1_493_313,
    },
}

MODEL_ARTIFACT_SPECS = {
    "first_estimates": {
        "path": "runs/first_estimates_v1.json",
        "sha256": (
            "719604ca4364e7cdef2293329ed0beb0e011e5d4d1c34f0e508c8f2fd9932977"
        ),
    },
    "anchor_context": {
        "path": "runs/anchor_context_report_v1.json",
        "sha256": (
            "1ff3f0a04d9a511079e7b1773e01de48a0f4a9b1ec405d4601d410eb88c2cb34"
        ),
    },
    "m2_pseudo_projection": {
        "path": "runs/m2_pseudo_projection_v1.json",
        "sha256": (
            "c55cf1ef34548fb75a7f27150f54e3b0b35db38ea3853ee71860e5d80964a715"
        ),
    },
    "replication_ppi_shared": {
        "path": "runs/replication_ppi_shared_v1.json",
        "sha256": (
            "36d6f9d1086d7884e11b2d828fa0ffb6ee7abe7940de881ebd477fd88f9630b5"
        ),
    },
}

MISSING_MODULE_DEVIATION = {
    "model_value": None,
    "status": "not_computable",
}


def sha256(raw: bytes) -> str:
    """Return the lowercase SHA-256 identity of bytes."""

    return hashlib.sha256(raw).hexdigest()


def verify_bytes(
    path: Path, expected_sha256: str, expected_size: int, description: str
) -> bytes:
    """Read one file and reject any byte-identity drift."""

    raw = path.read_bytes()
    if len(raw) != expected_size or sha256(raw) != expected_sha256:
        raise AssertionError(f"{description} identity drifted: {path}")
    return raw


def verified_inputs() -> dict[str, Any]:
    """Verify every capture and model artifact before parsing their bytes."""

    manifest_raw = verify_bytes(
        VALIDATION_MANIFEST,
        VALIDATION_MANIFEST_SHA256,
        VALIDATION_MANIFEST_SIZE_BYTES,
        "validation-source manifest",
    )
    manifest_entries = [
        json.loads(line)
        for line in manifest_raw.decode("utf-8").splitlines()
        if line
    ]
    if len(manifest_entries) != 30:
        raise AssertionError("validation-source manifest entry count drifted")
    by_filename = {entry["filename"]: entry for entry in manifest_entries}
    if len(by_filename) != len(manifest_entries):
        raise AssertionError("validation-source manifest filenames collide")

    capture_raw: dict[str, bytes] = {}
    capture_identities: dict[str, dict[str, Any]] = {}
    for capture_id, expected in CAPTURE_SPECS.items():
        filename = expected["filename"]
        entry = by_filename.get(filename)
        if entry is None:
            raise AssertionError(
                f"capture is absent from manifest: {filename}"
            )
        if (
            entry["sha256"] != expected["sha256"]
            or entry["size_bytes"] != expected["size_bytes"]
        ):
            raise AssertionError(
                f"manifested capture identity drifted: {filename}"
            )
        capture_raw[capture_id] = verify_bytes(
            VALIDATION_SOURCES / filename,
            expected["sha256"],
            expected["size_bytes"],
            f"audited capture {capture_id}",
        )
        capture_identities[capture_id] = deepcopy(entry)

    ssa_manifest_raw = verify_bytes(
        SSA_CAPTURE_MANIFEST,
        SSA_CAPTURE_MANIFEST_SHA256,
        SSA_CAPTURE_MANIFEST_SIZE_BYTES,
        "SSA snapshot capture manifest",
    )
    ssa_capture_raw = verify_bytes(
        SSA_CAPTURE,
        SSA_CAPTURE_SHA256,
        SSA_CAPTURE_SIZE_BYTES,
        "SSA Supplement 4.B capture",
    )
    matching_lines = [
        line
        for line in ssa_manifest_raw.decode("utf-8").splitlines()
        if line.endswith(" supplement2025_4b.html")
    ]
    if len(matching_lines) != 1:
        raise AssertionError("SSA 4.B snapshot manifest entry is ambiguous")
    if matching_lines[0].split()[1:] != [
        SSA_CAPTURE_SHA256,
        str(SSA_CAPTURE_SIZE_BYTES),
        "supplement2025_4b.html",
    ]:
        raise AssertionError("SSA 4.B snapshot manifest entry drifted")
    ssa_extraction_raw = verify_bytes(
        SSA_EXTRACTION,
        SSA_EXTRACTION_SHA256,
        SSA_EXTRACTION_SIZE_BYTES,
        "committed SSA 4.B7 extraction",
    )
    capture_identities["ssa_supplement_4b"] = {
        "filename": SSA_CAPTURE.relative_to(ROOT).as_posix(),
        "sha256": SSA_CAPTURE_SHA256,
        "size_bytes": SSA_CAPTURE_SIZE_BYTES,
        "snapshot_manifest": {
            "path": SSA_CAPTURE_MANIFEST.relative_to(ROOT).as_posix(),
            "sha256": SSA_CAPTURE_MANIFEST_SHA256,
            "size_bytes": SSA_CAPTURE_MANIFEST_SIZE_BYTES,
        },
        "url": (
            "https://www.ssa.gov/policy/docs/statcomps/supplement/2025/"
            "4b.html#table4.b7"
        ),
    }

    model_raw: dict[str, bytes] = {}
    for artifact_id, expected in MODEL_ARTIFACT_SPECS.items():
        path = ROOT / expected["path"]
        raw = path.read_bytes()
        if sha256(raw) != expected["sha256"]:
            raise AssertionError(f"model artifact identity drifted: {path}")
        model_raw[artifact_id] = raw

    return {
        "capture_identities": capture_identities,
        "capture_raw": capture_raw,
        "manifest_entries": manifest_entries,
        "manifest_raw": manifest_raw,
        "model_raw": model_raw,
        "ssa_capture_raw": ssa_capture_raw,
        "ssa_extraction_raw": ssa_extraction_raw,
    }


def artifact_pointer(artifact_id: str, json_pointer: str) -> dict[str, str]:
    """Return one immutable repository artifact pointer."""

    artifact = MODEL_ARTIFACT_SPECS[artifact_id]
    return {
        "json_pointer": json_pointer,
        "path": artifact["path"],
        "sha256": artifact["sha256"],
    }


def missing_ours(
    unit: str,
    pointer: dict[str, str],
    required_module: str,
) -> dict[str, Any]:
    """Return one honest missing-module model-side record."""

    return {
        "artifact_pointer": pointer,
        "required_module": required_module,
        "status": "module_missing",
        "unit": unit,
        "value": None,
    }


def row_action(
    *,
    action: str,
    row_id: str,
    source_family: str,
    tier: str,
    published_unit: str,
    published_value: list[dict[str, Any]],
    source: dict[str, Any],
    ours: dict[str, Any],
    deviation: dict[str, Any],
) -> dict[str, Any]:
    """Assemble one inventory action in the evaluation-artifact schema."""

    return {
        "action": action,
        "deviation": deviation,
        "gap_class": (
            "module_missing"
            if ours["status"] == "module_missing"
            else "concept_mismatch"
        ),
        "our": ours,
        "published": {
            "unit": published_unit,
            "value": published_value,
        },
        "row_id": row_id,
        "source": source,
        "source_family": source_family,
        "tier": tier,
    }


def finite_number(value: Any, locator: str) -> float | int:
    """Require one finite, nonboolean numeric workbook value."""

    if (
        isinstance(value, bool)
        or not isinstance(value, (int, float))
        or not math.isfinite(value)
    ):
        raise AssertionError(f"expected finite numeric cell at {locator}")
    return value


def xlsx(raw: bytes):
    """Open verified XLSX bytes without following a mutable source path."""

    return load_workbook(BytesIO(raw), read_only=True, data_only=True)


def cbo_source(
    capture_id: str,
    *,
    sheet: str,
    locator: str,
    transform: str,
) -> dict[str, Any]:
    """Return one compact CBO capture locator."""

    return {
        "capture_id": capture_id,
        "locator": {"cells": locator, "sheet": sheet},
        "transform": transform,
    }


def build_initial_replacement_rows(raw: bytes) -> list[dict[str, Any]]:
    """Extract the 12 legitimate CBO 55038 replacement-rate paths."""

    workbook = xlsx(raw)
    sheet = workbook["Exhibit 5"]
    specs = (
        ("all.last5.scheduled", 14, 2),
        ("all.last5.payable", 14, 3),
        ("all.age22_61.scheduled", 14, 5),
        ("all.age22_61.payable", 14, 6),
        ("q1.last5.scheduled", 20, 2),
        ("q1.last5.payable", 20, 3),
        ("q1.age22_61.scheduled", 20, 5),
        ("q1.age22_61.payable", 20, 6),
        ("q5.last5.scheduled", 26, 2),
        ("q5.last5.payable", 26, 3),
        ("q5.age22_61.scheduled", 26, 5),
        ("q5.age22_61.payable", 26, 6),
    )
    rows = []
    for suffix, first_row, column in specs:
        observations = []
        for row_number in range(first_row, first_row + 3):
            cohort = sheet.cell(row_number, 1).value
            if cohort not in {"1940s", "1960s", "1980s"}:
                raise AssertionError("CBO 55038 cohort locator drifted")
            value = finite_number(
                sheet.cell(row_number, column).value,
                f"Exhibit 5!{row_number},{column}",
            )
            observations.append({"birth_cohort": cohort, "percent": value})
        row_id = "cbo.55038.initial_replacement_rate." f"{suffix}.cohort_path"
        rows.append(
            row_action(
                action="add",
                row_id=row_id,
                source_family="cbo",
                tier="model_triangulation",
                published_unit="percent",
                published_value=observations,
                source=cbo_source(
                    "cbo_55038_supplemental",
                    sheet="Exhibit 5",
                    locator=(
                        f"B{first_row - 2}; B{first_row - 1}; "
                        f"A{first_row}:A{first_row + 2}+"
                        f"{sheet.cell(first_row, column).coordinate}:"
                        f"{sheet.cell(first_row + 2, column).coordinate}"
                    ),
                    transform="identity_percent",
                ),
                ours=missing_ours(
                    "percent",
                    artifact_pointer(
                        "replication_ppi_shared",
                        "/full_sample/by_shared_quintile",
                    ),
                    (
                        "cohort-by-shared-lifetime-earnings-quintile initial "
                        "replacement rates with matched earnings denominators "
                        "and scheduled/payable scenarios"
                    ),
                ),
                deviation=deepcopy(MISSING_MODULE_DEVIATION),
            )
        )
    workbook.close()
    return rows


def build_lifetime_ratio_rows(raw: bytes) -> list[dict[str, Any]]:
    """Extract the 12 CBO 60392 lifetime benefit-to-tax paths."""

    workbook = xlsx(raw)
    sheet = workbook["14"]
    group_rows = (
        ("all", 12),
        ("q1", 18),
        ("q2", 24),
        ("q3", 30),
        ("q4", 36),
        ("q5", 42),
    )
    rows = []
    for group, first_row in group_rows:
        for scenario, column in (("scheduled", 2), ("payable", 3)):
            observations = []
            for row_number in range(first_row, first_row + 5):
                cohort = sheet.cell(row_number, 1).value
                if cohort not in {
                    "1950s",
                    "1960s",
                    "1970s",
                    "1980s",
                    "1990s",
                }:
                    raise AssertionError("CBO 60392 cohort locator drifted")
                value = finite_number(
                    sheet.cell(row_number, column).value,
                    f"14!{row_number},{column}",
                )
                observations.append({"birth_cohort": cohort, "ratio": value})
            row_id = (
                "cbo.60392.lifetime_benefit_tax_ratio."
                f"{group}.{scenario}.cohort_path"
            )
            rows.append(
                row_action(
                    action="add",
                    row_id=row_id,
                    source_family="cbo",
                    tier="model_triangulation",
                    published_unit=(
                        "dimensionless lifetime benefit-to-tax ratio"
                    ),
                    published_value=observations,
                    source=cbo_source(
                        "cbo_60392_long_term",
                        sheet="14",
                        locator=(
                            f"B{first_row - 1}; "
                            f"A{first_row}:A{first_row + 4}+"
                            f"{sheet.cell(first_row, column).coordinate}:"
                            f"{sheet.cell(first_row + 4, column).coordinate}"
                        ),
                        transform="identity_dimensionless_ratio",
                    ),
                    ours=missing_ours(
                        "dimensionless lifetime benefit-to-tax ratio",
                        artifact_pointer(
                            "m2_pseudo_projection", "/balance_analogue"
                        ),
                        (
                            "person-level lifetime benefit and payroll-tax "
                            "present values with age-65 discounting, benefit-"
                            "income-tax netting, household quintiles, and "
                            "payable vintages"
                        ),
                    ),
                    deviation=deepcopy(MISSING_MODULE_DEVIATION),
                )
            )
    workbook.close()
    return rows


def annual_percent_path(sheet: Any, column: int) -> list[dict[str, Any]]:
    """Extract the audited 2015-2022 and 2024-2098 annual CBO path."""

    row_numbers = [*range(40, 48), *range(49, 124)]
    expected_years = [*range(2015, 2023), *range(2024, 2099)]
    observations = []
    for row_number, expected_year in zip(
        row_numbers, expected_years, strict=True
    ):
        year = sheet.cell(row_number, 1).value
        if year != expected_year:
            raise AssertionError("CBO annual path year locator drifted")
        value = finite_number(
            sheet.cell(row_number, column).value,
            f"{sheet.title}!{row_number},{column}",
        )
        observations.append({"percent": value, "year": year})
    return observations


def model_tax_path(first: dict[str, Any]) -> list[dict[str, Any]]:
    """Derive the existing overlap-only tax-share model value."""

    rows = first["tables"]["revenue"]["per_draw"]
    by_year: dict[int, list[float]] = defaultdict(list)
    for item in rows:
        by_year[item["year"]].append(
            100.0
            * item["combined_contributions"]
            / item["weighted_taxable_payroll"]
        )
    observations = []
    for year in range(2015, 2023):
        values = by_year[year]
        if not values or max(values) - min(values) >= 1e-12:
            raise AssertionError(
                "model tax share is not a mechanical identity"
            )
        mean = statistics.mean(values)
        if abs(mean - 12.4) >= 1e-12:
            raise AssertionError("model tax share changed from 12.4 percent")
        observations.append(
            {
                "mean_percent": 12.4,
                "sample_sd_across_draws": 0.0,
                "year": year,
            }
        )
    return observations


def model_outlay_path(anchor: dict[str, Any]) -> list[dict[str, Any]]:
    """Read the committed overlap-only benefit/payroll proxy path."""

    comparison = anchor["results"]["comparison_results"][8]
    if (
        comparison["comparison_id"]
        != "cmp_retired_worker_benefits_per_reported_taxable_earnings"
        or comparison["availability"] != "available"
        or comparison["evaluated"] is not True
    ):
        raise AssertionError("anchor-context outlay proxy identity drifted")
    observations = []
    for expected_year, item in zip(
        range(2015, 2023), comparison["annual_rows"], strict=True
    ):
        if item["year"] != expected_year:
            raise AssertionError("anchor-context outlay proxy years drifted")
        observations.append(
            {
                "mean_percent": 100.0 * item["model_statistic_mean"],
                "sample_sd_across_draws": (
                    100.0 * item["model_statistic_sample_sd"]
                ),
                "year": expected_year,
            }
        )
    return observations


def percentage_point_deviation(
    ours: list[dict[str, Any]], published: list[dict[str, Any]]
) -> dict[str, Any]:
    """Compare two annual percent paths on their shared years."""

    published_by_year = {item["year"]: item["percent"] for item in published}
    return {
        "definition": "ours minus published, in percentage points",
        "signed_percentage_points": [
            {
                "percentage_points": (
                    item["mean_percent"] - published_by_year[item["year"]]
                ),
                "year": item["year"],
            }
            for item in ours
        ],
    }


def indexed_path(sheet: Any, *, total_column: int = 4) -> list[dict[str, Any]]:
    """Index one 2024-2098 CBO total path to its first observation."""

    base = finite_number(sheet.cell(9, total_column).value, "row 9 total")
    observations = []
    for row_number, expected_year in zip(
        range(9, 84), range(2024, 2099), strict=True
    ):
        year = sheet.cell(row_number, 1).value
        if year != expected_year:
            raise AssertionError("CBO indexed path year locator drifted")
        value = finite_number(
            sheet.cell(row_number, total_column).value,
            f"{sheet.title}!{row_number},{total_column}",
        )
        observations.append(
            {"index_2024_100": 100.0 * value / base, "year": year}
        )
    return observations


def trust_fund_path(sheet: Any) -> list[dict[str, Any]]:
    """Extract all audited trust-fund ratio cells, retaining printed n.a."""

    row_numbers = [*range(40, 48), *range(49, 91)]
    expected_years = [*range(2015, 2023), *range(2024, 2066)]
    observations = []
    for row_number, expected_year in zip(
        row_numbers, expected_years, strict=True
    ):
        year = sheet.cell(row_number, 1).value
        if year != expected_year:
            raise AssertionError("CBO trust-fund path year locator drifted")
        observation: dict[str, Any] = {"year": year}
        for key, column in (
            ("oasi_ratio", 2),
            ("di_ratio", 3),
            ("oasdi_ratio", 4),
        ):
            value = sheet.cell(row_number, column).value
            if value == "n.a.":
                observation[key] = None
            else:
                observation[key] = finite_number(
                    value, f"8!{row_number},{column}"
                )
        observations.append(observation)
    return observations


def build_cbo_annual_rows(
    long_term_raw: bytes,
    additional_raw: bytes,
    first: dict[str, Any],
    anchor: dict[str, Any],
) -> list[dict[str, Any]]:
    """Build six CBO annual actions, including one legacy-row revision."""

    long_term = xlsx(long_term_raw)
    sheet_1 = long_term["1"]
    sheet_3 = long_term["3"]
    sheet_8 = long_term["8"]
    published_outlays_payroll = annual_percent_path(sheet_1, 3)
    published_tax_payroll = annual_percent_path(sheet_1, 2)
    published_outlays_gdp = annual_percent_path(sheet_3, 3)
    our_outlays = model_outlay_path(anchor)
    our_tax = model_tax_path(first)

    additional = xlsx(additional_raw)
    beneficiary_index = indexed_path(additional["3. OASDI Beneficiaries"])
    worker_index = indexed_path(additional["1. Covered Workers"])
    trust_path = trust_fund_path(sheet_8)

    rows = [
        row_action(
            action="add",
            row_id="cbo.60392.oasdi_outlays_share_taxable_payroll.path",
            source_family="cbo",
            tier="model_triangulation",
            published_unit="percent of national taxable payroll",
            published_value=published_outlays_payroll,
            source=cbo_source(
                "cbo_60392_long_term",
                sheet="1",
                locator=(
                    "A9:D9 headers; A40:A47+C40:C47; "
                    "A49:A123+C49:C123; notes A127, A129"
                ),
                transform="identity_percent",
            ),
            ours={
                "artifact_pointer": artifact_pointer(
                    "anchor_context",
                    "/results/comparison_results/8/annual_rows",
                ),
                "formula": "100 * model_statistic_mean",
                "required_module": (
                    "projection-vintage full-OASDI outlay path"
                ),
                "status": "available_proxy_overlap_only",
                "unit": "percent of frame-relative proxy taxable payroll",
                "value": our_outlays,
            },
            deviation=percentage_point_deviation(
                our_outlays, published_outlays_payroll
            ),
        ),
        row_action(
            action="revise",
            row_id="cbo.tax_revenue.share_of_taxable_payroll",
            source_family="cbo",
            tier="model_triangulation",
            published_unit="percent of national taxable payroll",
            published_value=published_tax_payroll,
            source=cbo_source(
                "cbo_60392_long_term",
                sheet="1",
                locator=(
                    "A9:D9 headers; A40:A47+B40:B47; "
                    "A49:A123+B49:B123; notes A127, A129"
                ),
                transform="identity_percent",
            ),
            ours={
                "artifact_pointer": artifact_pointer(
                    "first_estimates", "/tables/revenue/per_draw"
                ),
                "formula": (
                    "100 * combined_contributions / "
                    "weighted_taxable_payroll"
                ),
                "required_module": (
                    "future tax-revenue path including benefit-income-tax "
                    "revenue"
                ),
                "status": "available_overlap_only",
                "unit": "percent of frame-relative proxy taxable payroll",
                "value": our_tax,
            },
            deviation=percentage_point_deviation(
                our_tax, published_tax_payroll
            ),
        ),
        row_action(
            action="add",
            row_id="cbo.60392.oasdi_outlays_share_gdp.path",
            source_family="cbo",
            tier="model_triangulation",
            published_unit="percent of GDP",
            published_value=published_outlays_gdp,
            source=cbo_source(
                "cbo_60392_long_term",
                sheet="3",
                locator=(
                    "A9:D9 headers; A40:A47+C40:C47; "
                    "A49:A123+C49:C123; notes A127, A129"
                ),
                transform="identity_percent",
            ),
            ours=missing_ours(
                "percent of GDP",
                artifact_pointer(
                    "anchor_context",
                    "/results/comparison_results/8/annual_rows",
                ),
                "GDP denominator plus full-OASDI outlay projection",
            ),
            deviation=deepcopy(MISSING_MODULE_DEVIATION),
        ),
        row_action(
            action="add",
            row_id="cbo.60392.oasdi_beneficiaries.trajectory_2024_100",
            source_family="cbo",
            tier="model_triangulation",
            published_unit="index, 2024 = 100",
            published_value=beneficiary_index,
            source=cbo_source(
                "cbo_60392_additional",
                sheet="3. OASDI Beneficiaries",
                locator="A8:D8 headers; A9:A83+D9:D83; notes A85:A87",
                transform="100 * total / 2024 total",
            ),
            ours=missing_ours(
                "index, 2024 = 100",
                artifact_pointer(
                    "anchor_context",
                    "/results/comparison_results/6/annual_rows",
                ),
                "projection-vintage all-OASDI beneficiary stock",
            ),
            deviation=deepcopy(MISSING_MODULE_DEVIATION),
        ),
        row_action(
            action="add",
            row_id="cbo.60392.covered_workers.trajectory_2024_100",
            source_family="cbo",
            tier="model_triangulation",
            published_unit="index, 2024 = 100",
            published_value=worker_index,
            source=cbo_source(
                "cbo_60392_additional",
                sheet="1. Covered Workers",
                locator="A8:D8 headers; A9:A83+D9:D83; notes A85:A89",
                transform="100 * total / 2024 total",
            ),
            ours=missing_ours(
                "index, 2024 = 100",
                artifact_pointer(
                    "first_estimates", "/tables/revenue/per_draw"
                ),
                "projection-vintage covered-worker population",
            ),
            deviation=deepcopy(MISSING_MODULE_DEVIATION),
        ),
        row_action(
            action="add",
            row_id="cbo.60392.trust_fund_ratios.path",
            source_family="cbo",
            tier="model_triangulation",
            published_unit="beginning-of-year asset-to-cost ratio",
            published_value=trust_path,
            source=cbo_source(
                "cbo_60392_long_term",
                sheet="8",
                locator=(
                    "A9:D9 headers; A40:A47+B40:D47; "
                    "A49:A90+B49:D90; notes A94, A96, A104, A106:A113"
                ),
                transform="identity_ratio; preserve printed n.a. as null",
            ),
            ours=missing_ours(
                "beginning-of-year asset-to-cost ratio",
                artifact_pointer(
                    "m2_pseudo_projection", "/exhaustion_analogue"
                ),
                (
                    "annual trust-fund balances and scheduled-payment "
                    "denominators by fund"
                ),
            ),
            deviation=deepcopy(MISSING_MODULE_DEVIATION),
        ),
    ]
    additional.close()
    long_term.close()
    return rows


class MintCellParser(HTMLParser):
    """Collect scoped MINT table cells with their exact headers attribute."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.div_stack: list[str | None] = []
        self.cells: dict[str, list[dict[str, str | None]]] = defaultdict(list)
        self.current_cell: dict[str, Any] | None = None

    @property
    def active_table_id(self) -> str | None:
        return next(
            (value for value in reversed(self.div_stack) if value), None
        )

    def handle_starttag(
        self, tag: str, attrs: list[tuple[str, str | None]]
    ) -> None:
        attributes = dict(attrs)
        if tag == "div":
            self.div_stack.append(attributes.get("id"))
        elif tag in {"th", "td"}:
            self.current_cell = {
                "headers": attributes.get("headers"),
                "parts": [],
                "table_id": self.active_table_id,
            }

    def handle_data(self, data: str) -> None:
        if self.current_cell is not None:
            self.current_cell["parts"].append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self.current_cell is not None:
            text = " ".join("".join(self.current_cell.pop("parts")).split())
            table_id = self.current_cell.pop("table_id")
            if table_id is not None:
                self.cells[table_id].append(
                    {**self.current_cell, "text": text}
                )
            self.current_cell = None
        elif tag == "div" and self.div_stack:
            self.div_stack.pop()


def mint_cell(parser: MintCellParser, table_id: str, headers: str) -> str:
    """Return one uniquely scoped MINT table cell."""

    matches = [
        cell["text"]
        for cell in parser.cells[table_id]
        if cell["headers"] == headers
    ]
    if len(matches) != 1 or not isinstance(matches[0], str):
        raise AssertionError(
            f"MINT selector did not resolve uniquely: {table_id} {headers}"
        )
    return matches[0]


def mint_number(text: str) -> float:
    """Parse one printed dollar/count MINT percentile."""

    cleaned = text.replace("$", "").replace(",", "").strip()
    try:
        value = float(cleaned)
    except ValueError as error:
        raise AssertionError(f"invalid MINT numeric cell: {text}") from error
    if not math.isfinite(value):
        raise AssertionError(f"nonfinite MINT numeric cell: {text}")
    return value


def build_mint_rows(
    beneficiary_raw: bytes, taxpayer_raw: bytes
) -> list[dict[str, Any]]:
    """Extract four beneficiary shares and 15 repeated cross-sections."""

    beneficiaries = MintCellParser()
    beneficiaries.feed(beneficiary_raw.decode("utf-8"))
    taxpayers = MintCellParser()
    taxpayers.feed(taxpayer_raw.decode("utf-8"))
    years = (2024, 2030, 2050, 2070)

    beneficiary_specs = (
        ("retired_worker_only", "r40"),
        ("widow_including_dual", "r41"),
        ("spousal_including_dual", "r42"),
        ("disabled_worker_only", "r43"),
    )
    rows = []
    for category, row_header in beneficiary_specs:
        observations = []
        for year in years:
            table_id = f"tableIncome{year}"
            headers = f"r1 r39 {row_header} c3 c7"
            text = mint_cell(beneficiaries, table_id, headers)
            if not text.endswith("%"):
                raise AssertionError("MINT income-share cell is not a percent")
            share = mint_number(text[:-1]) / 100.0
            observations.append({"share": share, "year": year})
        row_id = (
            "ssa.mint8.beneficiaries."
            f"{category}.social_security_income_share.trajectory"
        )
        rows.append(
            row_action(
                action="add",
                row_id=row_id,
                source_family="mint",
                tier="model_triangulation",
                published_unit="fraction of household income",
                published_value=observations,
                source={
                    "capture_id": "mint_beneficiaries",
                    "locator": {
                        "selector": (
                            "div#tableIncome{YEAR} > table "
                            f"td[headers='r1 r39 {row_header} c3 c7']"
                        ),
                        "tables": "zero-based ordinals 8-11",
                    },
                    "transform": "parse printed percent and divide by 100",
                },
                ours=missing_ours(
                    "fraction of household income",
                    artifact_pointer(
                        "first_estimates",
                        "/tables/modeled_award_flow/per_draw",
                    ),
                    (
                        "beneficiary-type stock plus household-income "
                        "denominator and source decomposition"
                    ),
                ),
                deviation=deepcopy(MISSING_MODULE_DEVIATION),
            )
        )

    metric_specs = (
        (
            "annual_tax",
            "Taxes",
            "/tables/revenue/per_draw",
            (
                "individual annual tax percentiles with taxable-maximum "
                "mechanics"
            ),
        ),
        (
            "covered_earnings",
            "Earnings",
            "/tables/revenue/per_draw",
            (
                "age-stratified uncapped individual covered-earnings "
                "percentiles"
            ),
        ),
        (
            "lifetime_qc",
            "Quarters",
            "/diagnostics",
            (
                "statutory quarters-of-coverage and projected lifetime-"
                "career percentiles"
            ),
        ),
    )
    age_specs = (
        ("31_39", "r15"),
        ("40_49", "r16"),
        ("50_59", "r17"),
        ("60_69", "r18"),
        ("70_plus", "r19"),
    )
    for metric, stem, pointer, required_module in metric_specs:
        for age, row_header in age_specs:
            observations = []
            for year in years:
                table_id = f"table{stem}{year}"
                values = []
                for column in ("c3", "c4", "c5"):
                    headers = f"r1 r14 {row_header} c2 {column}"
                    values.append(
                        mint_number(mint_cell(taxpayers, table_id, headers))
                    )
                p10, median, p90 = values
                if median == 0:
                    raise AssertionError("MINT median cannot be zero")
                observations.append(
                    {
                        "p10_to_median": p10 / median,
                        "p90_to_median": p90 / median,
                        "year": year,
                    }
                )
            row_id = (
                f"ssa.mint8.taxpayers.age_{age}.{metric}_percentile_shape."
                "repeated_cross_section"
            )
            rows.append(
                row_action(
                    action="add",
                    row_id=row_id,
                    source_family="mint",
                    tier="model_triangulation",
                    published_unit=(
                        "within-age p10/median and p90/median ratios"
                    ),
                    published_value=observations,
                    source={
                        "capture_id": "mint_taxpayers",
                        "locator": {
                            "panel": "All taxpayers r1 > Age r14",
                            "selectors": {
                                name: (
                                    f"div#table{stem}{{YEAR}} > table "
                                    "td[headers='r1 r14 "
                                    f"{row_header} c2 {column}']"
                                )
                                for name, column in (
                                    ("p10", "c3"),
                                    ("median", "c4"),
                                    ("p90", "c5"),
                                )
                            },
                        },
                        "transform": {
                            "lower_tail": "p10 / median",
                            "upper_tail": "p90 / median",
                        },
                    },
                    ours=missing_ours(
                        "within-age p10/median and p90/median ratios",
                        artifact_pointer("first_estimates", pointer),
                        required_module,
                    ),
                    deviation=deepcopy(MISSING_MODULE_DEVIATION),
                )
            )
    return rows


SSA_ROW_ORDER = (
    "ssa_4b7_all_workers_share_1_9999",
    "ssa_4b7_all_workers_share_10000_19999",
    "ssa_4b7_all_workers_share_20000_39999",
    "ssa_4b7_all_workers_share_40000_59999",
    "ssa_4b7_all_workers_share_60000_79999",
    "ssa_4b7_all_workers_share_80000_99999",
    "ssa_4b7_all_workers_share_100000_119999",
    "ssa_4b7_all_workers_share_120000_139999",
    "ssa_4b7_all_workers_share_140000_149999",
    "ssa_4b7_all_workers_share_150000_160199",
    "ssa_4b7_all_workers_share_at_taxable_maximum",
)


def build_ssa_rows(extraction_raw: bytes) -> list[dict[str, Any]]:
    """Read the committed hash-bound Table 4.B7 same-panel extraction."""

    extraction = json.loads(extraction_raw)
    if (
        extraction["schema_version"] != "ssa_supplement_4b7_band_shares.v1"
        or extraction["source"]["capture_sha256"] != SSA_CAPTURE_SHA256
        or extraction["source"]["locator"] != "div#table4.b7 > table"
    ):
        raise AssertionError("SSA 4.B7 committed extraction metadata drifted")
    rows = []
    for row_id in SSA_ROW_ORDER:
        source_row = extraction["rows"][row_id]
        observations = [
            {"share": item["share"], "year": item["year"]}
            for item in source_row["observations"]
        ]
        rows.append(
            row_action(
                action="add",
                row_id=row_id,
                source_family="ssa_supplement_4b7",
                tier="admin_truth",
                published_unit="fraction of same-year all-worker total",
                published_value=observations,
                source={
                    "capture_id": "ssa_supplement_4b",
                    "committed_extraction": {
                        "json_pointer": f"/rows/{row_id}/observations",
                        "path": SSA_EXTRACTION.relative_to(ROOT).as_posix(),
                        "sha256": SSA_EXTRACTION_SHA256,
                    },
                    "locator": {
                        "denominator": "same-panel all-worker Total",
                        "header": source_row["header"],
                        "one_based_data_column": source_row[
                            "one_based_data_column"
                        ],
                        "table": "div#table4.b7 > table",
                    },
                    "transform": (
                        "printed band count / same-year same-panel Total; "
                        "discard structural N/A cells"
                    ),
                },
                ours=missing_ours(
                    "fraction of same-year all-worker total",
                    artifact_pointer(
                        "first_estimates", "/tables/revenue/per_draw"
                    ),
                    (
                        "weighted nominal taxable-earnings bands, wage-and-"
                        "salary membership, sex, and at-maximum indicator"
                    ),
                ),
                deviation=deepcopy(MISSING_MODULE_DEVIATION),
            )
        )
    return rows


def validate_rows(rows: list[dict[str, Any]]) -> None:
    """Fail closed on the resolved tranche inventory and sentinel census."""

    row_ids = [row["row_id"] for row in rows]
    if len(rows) != 60 or len(set(row_ids)) != 60:
        raise AssertionError("tranche must contain 60 unique row actions")
    if Counter(row["action"] for row in rows) != {"add": 59, "revise": 1}:
        raise AssertionError("tranche action census drifted")
    revised = [row["row_id"] for row in rows if row["action"] == "revise"]
    if revised != ["cbo.tax_revenue.share_of_taxable_payroll"]:
        raise AssertionError("legacy CBO tax row revision identity drifted")
    if Counter(row["source_family"] for row in rows) != {
        "cbo": 30,
        "mint": 19,
        "ssa_supplement_4b7": 11,
    }:
        raise AssertionError("tranche source census drifted")
    if Counter(row["tier"] for row in rows) != {
        "admin_truth": 11,
        "model_triangulation": 49,
    }:
        raise AssertionError("tranche tier census drifted")
    available = [row for row in rows if row["our"]["value"] is not None]
    missing = [row for row in rows if row["our"]["value"] is None]
    if len(available) != 2 or len(missing) != 58:
        raise AssertionError("tranche model-value census drifted")
    expected_available = {
        "cbo.tax_revenue.share_of_taxable_payroll",
        "cbo.60392.oasdi_outlays_share_taxable_payroll.path",
    }
    if {row["row_id"] for row in available} != expected_available:
        raise AssertionError("unexpected tranche model measurement")
    for row in missing:
        if (
            row["gap_class"] != "module_missing"
            or row["our"]["status"] != "module_missing"
            or row["deviation"] != MISSING_MODULE_DEVIATION
        ):
            raise AssertionError(
                f"dishonest missing-module sentinel: {row['row_id']}"
            )
    if any(
        row_id.startswith("cbo.55038.lifetime_benefit_tax_ratio")
        for row_id in row_ids
    ):
        raise AssertionError("CBO 55038 cannot supply benefit-to-tax ratios")
    forbidden_mint_fragments = (
        "beneficiary_type_population_share",
        "dual_entitlement_rate",
        "birth_cohort",
    )
    if any(
        row_id.startswith("ssa.mint8.")
        and any(fragment in row_id for fragment in forbidden_mint_fragments)
        for row_id in row_ids
    ):
        raise AssertionError("unsupported MINT denominator was invented")


def build() -> dict[str, Any]:
    """Build the canonical immutable evaluation payload."""

    inputs = verified_inputs()
    captures = inputs["capture_raw"]
    models = {
        artifact_id: json.loads(raw)
        for artifact_id, raw in inputs["model_raw"].items()
    }
    rows = [
        *build_initial_replacement_rows(captures["cbo_55038_supplemental"]),
        *build_lifetime_ratio_rows(captures["cbo_60392_long_term"]),
        *build_cbo_annual_rows(
            captures["cbo_60392_long_term"],
            captures["cbo_60392_additional"],
            models["first_estimates"],
            models["anchor_context"],
        ),
        *build_mint_rows(
            captures["mint_beneficiaries"],
            captures["mint_taxpayers"],
        ),
        *build_ssa_rows(inputs["ssa_extraction_raw"]),
    ]
    validate_rows(rows)
    return {
        "capture_identities": inputs["capture_identities"],
        "inventory": {
            "action_counts": {"add": 59, "revise": 1},
            "audited_capture_count": 7,
            "model_value_count": 2,
            "null_module_missing_count": 58,
            "row_action_count": 60,
            "source_counts": {
                "cbo": 30,
                "mint": 19,
                "ssa_supplement_4b7": 11,
            },
            "tier_counts": {
                "admin_truth": 11,
                "model_triangulation": 49,
            },
        },
        "model_artifacts": deepcopy(MODEL_ARTIFACT_SPECS),
        "rows": rows,
        "schema_version": "benchmark_tranche2_evaluation.v1",
        "standing_corrections": [
            (
                "CBO 55038 has no lifetime benefit-to-tax ratios; its present-"
                "value material is replacement-rate material, and the ratio "
                "paths come from CBO 60392 Long-Term sheet 14."
            ),
            (
                "The captured MINT pages identify neither beneficiary-type "
                "population shares, dual-entitlement rates, nor taxpayer "
                "birth-cohort distributions, so no such denominators are "
                "registered."
            ),
        ],
        "validation_only": True,
        "validation_source_manifest": {
            "entry_count": 30,
            "external_path": (
                "~/PolicyEngine/psid-data/validation-sources/manifest.jsonl"
            ),
            "sha256": VALIDATION_MANIFEST_SHA256,
            "size_bytes": VALIDATION_MANIFEST_SIZE_BYTES,
        },
    }


def render() -> bytes:
    """Render canonical sorted JSON bytes."""

    return (
        json.dumps(
            build(),
            sort_keys=True,
            indent=2,
            ensure_ascii=True,
            allow_nan=False,
        )
        + "\n"
    ).encode("utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check",
        action="store_true",
        help="fail if the committed artifact differs; never write",
    )
    args = parser.parse_args()
    expected = render()
    if args.check:
        if not OUT.exists() or OUT.read_bytes() != expected:
            raise SystemExit(f"generated artifact is stale: {OUT}")
        return
    OUT.write_bytes(expected)


if __name__ == "__main__":
    main()
